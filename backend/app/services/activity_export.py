"""Excel (.xlsx) export per activiteit-onderdeel (#85).

Genereert een werkboek met, per onderdeel (ActivitySubRegistration), één rij per
inschrijving: aantallen per product + financials (verschuldigd, betaald
online/overschrijving-cash, terugbetaald, saldo) zoals ze NU in de DB staan, met
een totaalrij. De live DB is de bron van waarheid — er wordt aangenomen dat de
penningmeester overschrijvingen via de admin heeft ingevoerd.

Bevat persoons- en financiële data: enkel admin, nooit in de repo.
"""
from decimal import Decimal
from io import BytesIO
from typing import List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.domains.payment_status.service import get_records_for
from app.services.registration_totals import compute_registration_total

_EURO_FMT = '#,##0.00'
_HEADER_FILL = PatternFill("solid", fgColor="E5E7EB")
_HEADER_FONT = Font(bold=True)
_TOTAL_FONT = Font(bold=True)
_TOP_BORDER = Border(top=Side(style="thin", color="9CA3AF"))

_METHOD_LABELS = {"ONLINE": "Online", "TRANSFER": "Overschrijving", "CASH": "Cash"}


def _registration_financials(db, reg) -> Tuple[Decimal, Decimal, Decimal, Decimal, Decimal]:
    """(verschuldigd, betaald_online, betaald_offline, terugbetaald, saldo)."""
    due, _ = compute_registration_total(reg)
    records = get_records_for(db, "registration", reg.id)
    paid_online = Decimal("0")
    paid_offline = Decimal("0")
    refunded = Decimal("0")
    for r in records:
        if r.amount_paid is None:
            continue
        amt = Decimal(str(r.amount_paid))
        if r.type == "refund":
            refunded += -amt  # refund-amount_paid is negatief → terugbetaald positief
        elif r.method == "online":
            paid_online += amt
        else:  # transfer / cash
            paid_offline += amt
    net = paid_online + paid_offline - refunded
    return due, paid_online, paid_offline, refunded, due - net


def _status_label(due: Decimal, saldo: Decimal) -> str:
    if saldo > Decimal("0.005"):
        return "Open"
    if saldo < Decimal("-0.005"):
        return "Te veel betaald"
    return "Vereffend" if due > 0 else "Gratis"


def build_component_export_xlsx(db, activity, component) -> bytes:
    """Bouw het .xlsx-werkboek voor één onderdeel en geef de bytes terug."""
    products = list(component.products)
    registrations = [r for r in activity.registrations if r.component_id == component.id]

    wb = Workbook()
    ws = wb.active
    ws.title = (component.name or "Onderdeel")[:31]  # Excel-tabbladlimiet

    headers = (
        ["Naam"]
        + [p.name for p in products]
        + ["Verschuldigd", "Betaald online", "Betaald overschr./cash",
           "Terugbetaald", "Saldo", "Betaalwijze", "Status"]
    )
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center" if col > 1 else "left")

    n_products = len(products)
    # Kolomindexen (1-based) van de geldkolommen, voor opmaak + totalen.
    first_money_col = 2 + n_products  # na Naam + productkolommen
    money_cols = list(range(first_money_col, first_money_col + 5))  # 5 geldkolommen

    product_totals = [0] * n_products
    money_totals = [Decimal("0")] * 5

    for reg in registrations:
        qty_by_product = {}
        for item in reg.items:
            qty_by_product[item.product_id] = qty_by_product.get(item.product_id, 0) + item.quantity
        due, online, offline, refunded, saldo = _registration_financials(db, reg)

        row = [reg.contact_name or "—"]
        for idx, p in enumerate(products):
            q = qty_by_product.get(p.id, 0)
            product_totals[idx] += q
            row.append(q)
        money_vals = [due, online, offline, refunded, saldo]
        for i, v in enumerate(money_vals):
            money_totals[i] += v
            row.append(float(v))
        row.append(_METHOD_LABELS.get(reg.payment_method or "", reg.payment_method or "—"))
        row.append(_status_label(due, saldo))
        ws.append(row)

    # Totaalrij
    total_row = ["Totaal"] + product_totals + [float(v) for v in money_totals] + ["", ""]
    ws.append(total_row)
    total_row_idx = ws.max_row
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=total_row_idx, column=col)
        cell.font = _TOTAL_FONT
        cell.border = _TOP_BORDER

    # Euro-opmaak op de geldkolommen (alle datarijen + totaalrij)
    for col in money_cols:
        for r in range(2, total_row_idx + 1):
            ws.cell(row=r, column=col).number_format = _EURO_FMT

    # Kolombreedtes globaal wat ruimer
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18 if col == 1 else 14
    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
