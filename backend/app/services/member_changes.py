"""Overzicht/export van alle ledendata-wijzigingen sinds een datum (#82).

Raak Nationaal heeft geen API; wijzigingen in dit portaal moeten manueel
overgetypt worden. Dit leest de append-only history-tabellen (recorded_at >=
since) en levert per wijziging een leesbare regel, zodat de admin ze één voor één
kan overnemen. Bevat persoonsdata: admin-only, nooit in de repo.
"""
from datetime import date, datetime, time, timezone
from io import BytesIO
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy.orm import Session

from app.models.history import (
    PersonHistory,
    MemberHistory,
    MemberPersonHistory,
    MembershipHistory,
    AddressHistory,
    ContactDetailHistory,
    PaymentRecordHistory,
    RegistrationItemHistory,
    ActivityHistory,
    ActivityDateHistory,
    ComponentHistory,
    ProductHistory,
)

_OPERATION_LABELS = {"insert": "Toegevoegd", "update": "Gewijzigd", "delete": "Verwijderd"}


def _fmt(value) -> str:
    return "" if value is None else str(value)


def _row(h, *, entity: str, entity_id: Optional[int], summary: str, group: str = "Leden") -> dict:
    return {
        "recorded_at": h.recorded_at,
        "group": group,
        "entity": entity,
        "entity_id": entity_id,
        "operation": h.operation,
        "operation_label": _OPERATION_LABELS.get(h.operation, h.operation),
        "action": h.action,
        "actor": h.actor,
        "summary": summary,
    }


def member_changes_since(db: Session, since: date) -> List[dict]:
    """Alle ledendata-wijzigingen met recorded_at >= since, nieuw → oud."""
    since_dt = datetime.combine(since, time.min, tzinfo=timezone.utc)
    rows: List[dict] = []

    for h in db.query(PersonHistory).filter(PersonHistory.recorded_at >= since_dt):
        naam = f"{_fmt(h.first_name)} {_fmt(h.last_name)}".strip() or "—"
        dob = f" (geb. {h.date_of_birth})" if h.date_of_birth else ""
        rows.append(_row(h, entity="Persoon", entity_id=h.person_id, summary=f"{naam}{dob}"))

    for h in db.query(MemberHistory).filter(MemberHistory.recorded_at >= since_dt):
        rows.append(_row(h, entity="Gezin", entity_id=h.member_id, summary=f"gezin #{_fmt(h.member_id)}"))

    for h in db.query(MemberPersonHistory).filter(MemberPersonHistory.recorded_at >= since_dt):
        rows.append(_row(
            h, entity="Gezinslid", entity_id=h.member_person_id,
            summary=f"persoon #{_fmt(h.person_id)} in gezin #{_fmt(h.member_id)} ({_fmt(h.relation_type)})",
        ))

    for h in db.query(AddressHistory).filter(AddressHistory.recorded_at >= since_dt):
        adres = f"{_fmt(h.street)} {_fmt(h.house_number)}".strip()
        if h.bus_number:
            adres += f" bus {h.bus_number}"
        rows.append(_row(
            h, entity="Adres", entity_id=h.address_id,
            summary=f"{adres} (persoon #{_fmt(h.person_id)}, postcode-id {_fmt(h.postal_code_id)})",
        ))

    for h in db.query(ContactDetailHistory).filter(ContactDetailHistory.recorded_at >= since_dt):
        rows.append(_row(
            h, entity="Contact", entity_id=h.contact_detail_id,
            summary=f"{_fmt(h.contact_type_code)}: {_fmt(h.value)} (persoon #{_fmt(h.person_id)})",
        ))

    for h in db.query(MembershipHistory).filter(MembershipHistory.recorded_at >= since_dt):
        rows.append(_row(
            h, entity="Lidmaatschap", entity_id=h.membership_id,
            summary=f"jaar {_fmt(h.year)}, actief={_fmt(h.is_active)}, {_fmt(h.valid_from)}–{_fmt(h.valid_to)} (gezin #{_fmt(h.member_id)})",
        ))

    rows.sort(key=lambda r: r["recorded_at"], reverse=True)
    return rows


# Objectgroepen voor de filter op de Wijzigingen-pagina (#189).
GROUPS = ["Leden", "Activiteiten", "Inschrijvingen", "Betalingen"]


def all_changes_since(
    db: Session, since: date, *, group: Optional[str] = None, actor: Optional[str] = None
) -> List[dict]:
    """Unified audit-feed (#189): alle history-tabellen sinds ``since``, met een
    objectgroep per rij; optioneel gefilterd op groep en/of actor. Nieuw → oud."""
    since_dt = datetime.combine(since, time.min, tzinfo=timezone.utc)
    rows: List[dict] = list(member_changes_since(db, since))  # groep "Leden"

    for h in db.query(ActivityHistory).filter(ActivityHistory.recorded_at >= since_dt):
        rows.append(_row(h, entity="Activiteit", entity_id=h.activity_id,
                         summary=_fmt(h.name) or f"activiteit #{_fmt(h.activity_id)}",
                         group="Activiteiten"))
    for h in db.query(ActivityDateHistory).filter(ActivityDateHistory.recorded_at >= since_dt):
        rows.append(_row(h, entity="Datum", entity_id=h.activity_date_id,
                         summary=f"{_fmt(h.start_date)}–{_fmt(h.end_date)} (activiteit #{_fmt(h.activity_id)})",
                         group="Activiteiten"))
    for h in db.query(ComponentHistory).filter(ComponentHistory.recorded_at >= since_dt):
        rows.append(_row(h, entity="Onderdeel", entity_id=h.component_id,
                         summary=f"{_fmt(h.name)} (activiteit #{_fmt(h.activity_id)})",
                         group="Activiteiten"))
    for h in db.query(ProductHistory).filter(ProductHistory.recorded_at >= since_dt):
        rows.append(_row(h, entity="Product", entity_id=h.product_id,
                         summary=f"{_fmt(h.name)} €{_fmt(h.price)} (onderdeel #{_fmt(h.component_id)})",
                         group="Activiteiten"))
    for h in db.query(RegistrationItemHistory).filter(RegistrationItemHistory.recorded_at >= since_dt):
        rows.append(_row(h, entity="Bestelregel", entity_id=h.registration_item_id,
                         summary=f"product #{_fmt(h.product_id)} ×{_fmt(h.quantity)} (inschrijving #{_fmt(h.registration_id)})",
                         group="Inschrijvingen"))
    for h in db.query(PaymentRecordHistory).filter(PaymentRecordHistory.recorded_at >= since_dt):
        rows.append(_row(h, entity="Betaling", entity_id=None,
                         summary=f"{_fmt(h.type)} €{_fmt(h.amount)} {_fmt(h.method)}/{_fmt(h.status)} ({_fmt(h.payable_type)} #{_fmt(h.payable_id)})",
                         group="Betalingen"))

    if group:
        rows = [r for r in rows if r["group"] == group]
    if actor:
        rows = [r for r in rows if (r["actor"] or "") == actor]
    rows.sort(key=lambda r: r["recorded_at"], reverse=True)
    return rows


def build_member_changes_xlsx(rows: List[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Ledenwijzigingen"
    headers = ["Tijdstip", "Wat", "Type", "ID", "Actie", "Door", "Details"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="E5E7EB")

    for r in rows:
        ts = r["recorded_at"]
        ts_str = ts.strftime("%Y-%m-%d %H:%M") if ts else ""
        ws.append([
            ts_str, r["operation_label"], r["entity"], r["entity_id"],
            r["action"], r["actor"] or "", r["summary"],
        ])

    widths = [16, 12, 14, 8, 26, 26, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
