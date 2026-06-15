"""Excel-export per onderdeel (#85): aantallen per product + financials uit de
live DB, met een totaalrij. Admin-only."""
from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook

from app.domains.payment_status.models import PaymentRecord
from tests.conftest import seed_activity_with_product

_XLSX_MIME = "spreadsheetml"


def test_export_requires_admin(client, db_session):
    _, comp, _ = seed_activity_with_product(db_session)
    resp = client.get(f"/api/v1/activities/{comp.activity_id}/components/{comp.id}/export")
    assert resp.status_code in (401, 403)


def test_export_unknown_component_404(client, db_session, admin_headers):
    _, comp, _ = seed_activity_with_product(db_session)
    resp = client.get(
        f"/api/v1/activities/{comp.activity_id}/components/{comp.id + 9999}/export",
        headers=admin_headers,
    )
    assert resp.status_code == 404


def test_export_quantities_and_financials(client, db_session, admin_headers):
    _, comp, product = seed_activity_with_product(db_session, price="18.00")
    activity_id = comp.activity_id

    # Inschrijving: 2 stuks → verschuldigd €36.
    reg_resp = client.post(f"/api/v1/activities/{activity_id}/register", json={
        "contact_name": "An Janssens", "contact_email": "an@example.com",
        "component_id": comp.id, "payment_method": "TRANSFER",
        "items": [{"product_id": product.id, "quantity": 2}],
    })
    assert reg_resp.status_code in (200, 201), reg_resp.text

    charge = db_session.query(PaymentRecord).filter(
        PaymentRecord.payable_type == "registration", PaymentRecord.type == "charge",
    ).order_by(PaymentRecord.created_at.desc()).first()
    # Penningmeester boekt de overschrijving (€36) en betaalt €6 terug.
    client.patch(f"/api/v1/payment-status/records/{charge.id}",
                 json={"status": "paid", "amount_paid": "36.00"}, headers=admin_headers)
    client.post(f"/api/v1/payment-status/records/{charge.id}/refund",
                json={"amount": "6.00"}, headers=admin_headers)

    resp = client.get(
        f"/api/v1/activities/{activity_id}/components/{comp.id}/export",
        headers=admin_headers,
    )
    assert resp.status_code == 200, resp.text
    assert _XLSX_MIME in resp.headers.get("content-type", "")
    assert "attachment" in resp.headers.get("content-disposition", "")

    wb = load_workbook(BytesIO(resp.content))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = list(rows[0])
    assert headers[0] == "Naam"
    assert "Verschuldigd" in headers
    assert "Terugbetaald" in headers

    i_due = headers.index("Verschuldigd")
    i_offline = headers.index("Betaald overschr./cash")
    i_refunded = headers.index("Terugbetaald")
    i_saldo = headers.index("Saldo")

    data = rows[1]
    assert data[0] == "An Janssens"
    assert data[1] == 2  # aantal van het product (eerste productkolom)
    assert data[i_due] == 36.0
    assert data[i_offline] == 36.0
    assert data[i_refunded] == 6.0
    assert data[i_saldo] == 6.0  # 36 verschuldigd - (36 betaald - 6 terug) = 6

    total = rows[-1]
    assert total[0] == "Totaal"
    assert total[1] == 2
    assert total[i_due] == 36.0
    assert total[i_refunded] == 6.0
    assert total[i_saldo] == 6.0
