"""Export/overzicht van ledendata-wijzigingen sinds een datum (#82)."""
from datetime import date, timedelta
from io import BytesIO

from openpyxl import load_workbook

from tests.conftest import seed_postal_code


def _family_payload(email="lid@example.com"):
    return {
        "street": "Milostraat", "house_number": "40", "postal_code": "2400",
        "payment_method": "transfer",
        "members": [{
            "last_name": "Janssens", "first_name": "An", "email": email,
            "mobile": "0470123456", "relation_type": "HOOFDLID",
        }],
    }


def _create_family(client, db_session):
    seed_postal_code(db_session)
    resp = client.post("/api/v1/families", json=_family_payload())
    assert resp.status_code == 201, resp.text


def test_member_changes_requires_admin(client):
    resp = client.get("/api/v1/admin/member-changes", params={"since": date.today().isoformat()})
    assert resp.status_code in (401, 403)


def test_member_changes_lists_recent_changes(client, db_session, admin_headers):
    _create_family(client, db_session)
    resp = client.get("/api/v1/admin/member-changes",
                      params={"since": date.today().isoformat()}, headers=admin_headers)
    assert resp.status_code == 200, resp.text
    rows = resp.json()
    entities = {r["entity"] for r in rows}
    # Een gezinsregistratie raakt minstens persoon, gezin, gezinslid en contact.
    assert {"Persoon", "Gezin", "Gezinslid"}.issubset(entities)
    person_row = next(r for r in rows if r["entity"] == "Persoon")
    assert person_row["operation_label"] == "Toegevoegd"
    assert "An" in person_row["summary"]


def test_member_changes_respects_since_date(client, db_session, admin_headers):
    _create_family(client, db_session)
    tomorrow = (date.today() + timedelta(days=1)).isoformat()
    resp = client.get("/api/v1/admin/member-changes",
                      params={"since": tomorrow}, headers=admin_headers)
    assert resp.status_code == 200
    assert resp.json() == []


def test_member_changes_xlsx_export(client, db_session, admin_headers):
    _create_family(client, db_session)
    resp = client.get("/api/v1/admin/member-changes/export",
                      params={"since": date.today().isoformat()}, headers=admin_headers)
    assert resp.status_code == 200, resp.text
    assert "spreadsheetml" in resp.headers.get("content-type", "")
    wb = load_workbook(BytesIO(resp.content))
    ws = wb.active
    headers = [c.value for c in ws[1]]
    assert headers[0] == "Tijdstip" and "Details" in headers
    assert ws.max_row >= 2  # minstens één wijziging
